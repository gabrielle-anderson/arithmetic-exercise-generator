import random

from openpyxl import *
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from constants import *

headerHeight = 2
sheetTitle = "Number Operation Work"
tinyColumnWidth= 5
smallColumnWidth = 10
largeColumnWidth = 20
headerFont = Font(name='Times New Roman', bold=True, size = 20)
problemFont = Font(name='Times New Roman', size = 16)

class ProblemSheetWriter:
    def writeStudentProblems(self, dateTime, name, studentSpecification, additionProblems, subtractionProblems, \
                             multiplicationProblems, divisionProblems):
        self.additionProblems = additionProblems
        self.subtractionProblems = subtractionProblems
        self.multiplicationProblems = multiplicationProblems
        self.divisionProblems = divisionProblems

        additionLevel = studentSpecification.additionLevel
        subtractionLevel = studentSpecification.subtractionLevel
        multiplicationLevel = studentSpecification.multiplicationLevel
        divisionLevel = studentSpecification.divisionLevel

        wb = Workbook()
        ws = wb.active
        self.ws = ws
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.title = name + "'s problems"

        problemsLevel = studentSpecification.toStringShort()
        sheetTitle = name

        ws.merge_cells('A1:B1')
        ws['A1'] = sheetTitle
        ws['A1'].font = headerFont

        ws.merge_cells('A2:C2')
        ws['A2'] = problemsLevel
        ws['A2'].font = headerFont

        ws.merge_cells('D1:F1')
        ws['D1'] = sheetTitle
        ws['D1'].font = headerFont

        ws.merge_cells('D2:F2')
        ws['D2'] = problemsLevel
        ws['D2'].font = headerFont

        ws.merge_cells('G1:I1')
        ws['G1'] = sheetTitle
        ws['G1'].font = headerFont

        ws.merge_cells('G2:I2')
        ws['G2'] = problemsLevel
        ws['G2'].font = headerFont

        ws.column_dimensions['A'].width = tinyColumnWidth
        ws.column_dimensions['B'].width = largeColumnWidth
        ws.column_dimensions['C'].width = smallColumnWidth

        ws.column_dimensions['D'].width = tinyColumnWidth
        ws.column_dimensions['E'].width = largeColumnWidth
        ws.column_dimensions['F'].width = smallColumnWidth

        ws.column_dimensions['G'].width = tinyColumnWidth
        ws.column_dimensions['H'].width = largeColumnWidth
        ws.column_dimensions['I'].width = smallColumnWidth

        self.writeProblems(name, studentSpecification, 1)
        self.writeProblems(name, studentSpecification, 4)
        self.writeProblems(name, studentSpecification, 7)

        if not os.path.exists(dateTime):
            os.makedirs(dateTime)

        fileName =  dateTime + os.sep + name + ".xlsx"
        wb.save(fileName)

    def writeProblems(self, name, studentSpecification, columnNo):
        problemList = studentSpecification.toProblemList()
        i = 1
        while (len(problemList) > 0):
            problemString = random.choice(problemList)
            problemList.remove(problemString)
            row = (i + headerHeight)
            problemLevel = studentSpecification.getProblemLevel(problemString)
            problemSpecifications = self.getProblemSpecification(problemString)
            assert problemLevel in problemSpecifications.keys(), \
                "Student %s has level %s for %s problems, but level %s has no specification." % \
                    (name, str(problemLevel), problemNames[problemString], str(problemLevel))

            problemSpecification = problemSpecifications[problemLevel]
            (x, y) = problemSpecification.randomOperands()

            self.ws.cell(row=row, column=(columnNo)).value = i
            self.ws.cell(row=row, column=(columnNo)).font = problemFont
            self.ws.cell(row=row, column=(columnNo + 1)).value = formatProblem(problemString, x.value(), y.value())
            self.ws.cell(row=row, column=(columnNo + 1)).alignment = Alignment(horizontal='left')
            self.ws.cell(row=row, column=(columnNo + 1)).font = problemFont
            self.ws.cell(row=row, column=(columnNo + 2)).value = "="
            self.ws.cell(row=row, column=(columnNo + 2)).alignment = Alignment(horizontal='left')
            self.ws.cell(row=row, column=(columnNo + 2)).font = problemFont

            i = i + 1

    def getProblemSpecification(self, problemName):
        if problemName == additionString:
            return self.additionProblems
        elif problemName == subtractionString:
            return self.subtractionProblems
        elif problemName == multiplicationString:
            return self.multiplicationProblems
        elif problemName == divisionString:
            return self.divisionProblems
        assert False, "Invalid problem string."
