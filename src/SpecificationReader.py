import openpyxl

from OperandSpecification import OperandSpecification


class SpecificationReader:
    def __init__(self):
        self.requiredSheetNames = ["Operand 1", "Operand 2", "Explicit number ranges", "Adjusting", "Misc"]
        self.workbook = None

    def getProblem(self, path, operation, problemSpecification):
        self.workbook = openpyxl.load_workbook(path)
        # print path
        # print self.workbook
        assert (self.workbook.get_sheet_names() == self.requiredSheetNames), \
            "Worksheet names should be %s and instead are %s." % (self.requiredSheetNames, self.workbook.get_sheet_names())



        firstOperandSheet = self.workbook.get_sheet_by_name("Operand 1")
        firstOperandSpecification = self.getOperandSpecification(firstOperandSheet)
        secondOperandSheet = self.workbook.get_sheet_by_name("Operand 2")
        secondOperandSpecification = self.getOperandSpecification(secondOperandSheet)


        explicitNumberRangesSheet = self.workbook.get_sheet_by_name("Explicit number ranges")
        firstOperandExplicitRange = self.getValuesInColumn(explicitNumberRangesSheet.iter_rows, 0)
        secondOperandExplicitRange = self.getValuesInColumn(explicitNumberRangesSheet.iter_rows, 1)
        resultExplicitRange = self.getValuesInColumn(explicitNumberRangesSheet.iter_rows, 2)

        firstOperandSpecification.explicitRange = firstOperandExplicitRange
        secondOperandSpecification.explicitRange = secondOperandExplicitRange

        adjustmentSheet = self.workbook.get_sheet_by_name("Adjusting")
        adjustments = {}
        for row in adjustmentSheet.iter_rows(row_offset=1):
            unit = row[0].value
            adjustment = row[1].value
            if (unit != None and adjustment != None):
                adjustments[float(unit)] = bool(adjustment)

        miscParametersSheet = self.workbook.get_sheet_by_name("Misc")
        miscParameters = {}
        for row in miscParametersSheet.iter_rows(row_offset=0):
            paramName = row[0].value
            paramValue = row[1].value
            if (paramName != None and paramValue!= None):
                miscParameters[paramName] = paramValue

        return problemSpecification(firstOperandSpecification, secondOperandSpecification, resultExplicitRange, \
                                    operation, adjustments, miscParameters)

    def getOperandSpecification(self, sheet):
        assert sheet['A1'].value == "Digit multiple of"
        assert sheet['B1'].value == "Lower bound"
        assert sheet['C1'].value == "Upper bound"

        operandSpecification = OperandSpecification()
        for row in sheet.iter_rows(row_offset=1):
            unit = row[0].value
            lowerBound = row[1].value
            upperBound = row[2].value
            if (unit != None and lowerBound != None and upperBound != None):
                operandSpecification.addDigit(float(unit), int(lowerBound), int(upperBound))
        return operandSpecification

    def getValuesInColumn(self, rows, columnNo):
        result = []
        for row in rows(row_offset=2):
            value = row[columnNo].value
            if value != None:
                result.append(value)
        return result